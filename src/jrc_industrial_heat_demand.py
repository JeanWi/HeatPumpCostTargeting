"""
Parse JRC-IDEES2023 'Industry' Excel files into a long table.

Outputs CSVs with columns: Country,Year,Sector,Type,Indicator,Value

Heuristics for hierarchy detection (in order):
 - openpyxl cell.alignment.indent
 - cell.font.bold
 - leading whitespace in the cell.value

Usage examples:
  python scripts/parse_jrc_idees_industry.py --input data/JRC-IDEES2023 --output output
  python scripts/parse_jrc_idees_industry.py --input data/JRC-IDEES2023 --country DE --output output

"""
import argparse
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.cell.cell import Cell

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

YEAR_RE = re.compile(r"^\s*(19|20)\d{2}\s*$")
NON_DIGIT_RE = re.compile(r"[^0-9.\-eE]")


def discover_countries(input_dir: Path) -> List[Path]:
    """Return subdirectories in input_dir excluding EU27."""
    dirs = [p for p in sorted(input_dir.iterdir()) if p.is_dir()]
    logger.info("Discovered %d country directories (excluded EU27)", len(dirs))
    return dirs


def get_cell_value(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int) -> Any:
    """Return the cell value, """
    return sheet.cell(row=row, column=col).value



def detect_header_row(sheet: openpyxl.worksheet.worksheet.Worksheet, max_scan_rows: int = 30) -> Optional[int]:
    """Scan the first rows to find the header row containing several year-like cells. Returns 1-based row index."""
    max_row = min(sheet.max_row, max_scan_rows)
    for r in range(1, max_row + 1):
        year_count = 0
        for c in range(1, sheet.max_column + 1):
            val = get_cell_value(sheet, r, c)
            if val is None:
                continue
            sval = str(val).strip()
            if YEAR_RE.match(sval):
                year_count += 1
            else:
                # sometimes the year is numeric (int/float)
                if isinstance(val, (int,)) and 1900 <= val <= 2100:
                    year_count += 1
        if year_count >= 2:
            logger.debug("Detected header row %d with %d year-like cells", r, year_count)
            return r
    logger.warning("No clear header row found in first %d rows; falling back to row 1", max_scan_rows)
    return 1


def detect_year_columns(sheet: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> Dict[int, int]:
    """Return mapping col_index -> year (int) for columns that look like year headers. cols are 1-based."""
    year_cols: Dict[int, int] = {}
    for c in range(1, sheet.max_column + 1):
        val = get_cell_value(sheet, header_row, c)
        if val is None:
            continue
        sval = str(val).strip()
        m = YEAR_RE.match(sval)
        if m:
            try:
                year_cols[c] = int(sval)
            except ValueError:
                pass
        else:
            if isinstance(val, int) and 1900 <= val <= 2100:
                year_cols[c] = int(val)
    logger.debug("Detected year columns: %s", year_cols)
    return year_cols


def normalize_label(text: Optional[str]) -> Optional[str]:
    if text is None:
        return None
    if not isinstance(text, str):
        text = str(text)
    # normalize whitespace and bullets
    text = text.replace('\xa0', ' ')
    text = text.strip()
    # remove leading bullets or dashes
    text = re.sub(r"^[\u2022\-\*\u25E6\s]+", "", text)
    # drop trailing colons
    text = re.sub(r":$", "", text)
    return text if text != "" else None


def infer_indent_score(cell: Cell, raw_value: Any) -> Optional[float]:
    """Return a numeric score for indentation/level detection. Higher score => deeper level (heuristic).

    Priority: cell.alignment.indent -> (0 if bold else None) -> leading spaces count -> None
    """
    try:
        align = cell.alignment
        if align is not None and getattr(align, 'indent', None) is not None:
            return float(align.indent)
    except Exception:
        pass
    try:
        if getattr(cell, 'font', None) is not None and getattr(cell.font, 'bold', False):
            # bold likely indicates a top-level heading -> low score
            return 0.0
    except Exception:
        pass
    if raw_value is None:
        return None
    s = str(raw_value)
    # count leading spaces/tabs
    leading = len(s) - len(s.lstrip(' \t'))
    if leading > 0:
        return float(leading)
    return None


def build_indent_mapping(scores: List[Optional[float]]) -> Dict[float, int]:
    """Map observed scores to compact levels (0..n). Missing (None) are not included in mapping.

    Returns mapping score -> level
    """
    uniq = sorted({s for s in scores if s is not None})
    mapping: Dict[float, int] = {}
    for i, s in enumerate(uniq):
        mapping[s] = i
    logger.debug("Indent mapping: %s", mapping)
    return mapping


def parse_numeric(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value)
    if s.strip() == "":
        return None
    # normalize spaces and non-breaking spaces
    s = s.replace('\xa0', '')
    s = s.replace(' ', '')
    # replace comma with dot (simple heuristic)
    s = s.replace(',', '.')
    # remove any character that's not digit, dot, minus or exponent
    s = NON_DIGIT_RE.sub('', s)
    try:
        return float(s)
    except Exception:
        return None


def process_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, country: str, sheet_type: str) -> List[Dict[str, Any]]:
    """Process a sheet and return list of long-form rows (dicts).

    Each dict: Country, Year, Sector, Type, Indicator, Value
    """
    logger.info("Processing sheet %s (country=%s, type=%s)", sheet.title, country, sheet_type)
    header_row = detect_header_row(sheet)
    year_cols = detect_year_columns(sheet, header_row)
    first_year_col = min(year_cols.keys())
    # Determine label/indicator columns: non-year columns before first_year_col
    label_cols = [c for c in range(1, first_year_col) if c <= sheet.max_column]
    sector_col = label_cols[0] if len(label_cols) >= 1 else 1
    indicator_col = label_cols[1] if len(label_cols) >= 2 else None

    # collect indent scores from sector_col for mapping
    scores: List[Optional[float]] = []
    for r in range(header_row + 1, sheet.max_row + 1):
        raw = get_cell_value(sheet, r, sector_col)
        cell = sheet.cell(row=r, column=sector_col)
        scores.append(infer_indent_score(cell, raw))
    indent_map = build_indent_mapping(scores)

    current_path: List[Optional[str]] = []
    rows_out: List[Dict[str, Any]] = []

    sector_short = sheet.title.split("_", 1)[0].strip()
    sector_long = get_cell_value(sheet, 1, 1).split(":", 1)[1].split("/", 1)[0].strip()

    for r in range(header_row + 1, sheet.max_row + 1):
        raw_sector = get_cell_value(sheet, r, sector_col)
        cell_sector = sheet.cell(row=r, column=sector_col)
        label = normalize_label(raw_sector)


        if label in ["Detailed split of energy consumption by subsector (ktoe)", "Detailed split of energy consumption (ktoe)"]:
            label = "Energy consumption (ktoe)"
        if label in ["Energy intensity (kgoe per t of output)", "Energy intensity (toe/physical output index)"]:
            label = "Energy intensity (toe/physical output index)"
        if label in ["Market shares of energy uses by subsector (%)", "Market shares of energy uses (%)"]:
            label = "Market shares of energy uses (%)"


        score = infer_indent_score(cell_sector, raw_sector)

        # get score from next row:
        sector_next = get_cell_value(sheet, r+1, sector_col)
        cell_next = sheet.cell(row=r+1, column=sector_col)
        score_next = infer_indent_score(cell_next, sector_next)

        if score is not None and score in indent_map:
            level = indent_map[score]
        elif score is None and len(current_path) > 0:
            # no new label info -> keep previous level
            level = len(current_path) - 1
        else:
            level = 0

        if label is not None:
            # ensure current_path has space
            if level >= len(current_path):
                # extend with None
                current_path.extend([None] * (level + 1 - len(current_path)))
            current_path[level] = label
            # truncate deeper levels
            current_path = current_path[: level + 1]

        # Build label string from current_path (skip None)
        label_parts = [p for p in current_path if p]
        label_str = " > ".join(label_parts) if label_parts else None
        if len(label_parts) >= 2:
            parent_label_str = label_str.split(" > ")[:-1]
        else:
            parent_label_str = ["Totals"]

        parent_label_str = " > ".join(parent_label_str)


        # Determine indicator
        indicator = None
        if indicator_col is not None:
            raw_ind = get_cell_value(sheet, r, indicator_col)
            indicator = normalize_label(raw_ind)
        if not indicator:
            # fallback to header of indicator_col if present
            if indicator_col is not None:
                hdr = get_cell_value(sheet, header_row, indicator_col)
                indicator = normalize_label(hdr)
        if not indicator:
            indicator = sheet.title

        if (sector_next is not None) and (score_next != 0) and (score_next > score):
            continue
        code = get_cell_value(sheet, r, 105)

        # For each year column, get value
        for col, year in year_cols.items():
            raw_val = get_cell_value(sheet, r, col)
            val = parse_numeric(raw_val)
            # Store row even if val is None so structure is preserved? We'll skip empty rows where sector and val missing
            if val is None:
                # nothing to record
                continue
            rowd = {
                "Country": country,
                "Year": year,
                "Sector_long": sector_long,
                "Sector_short": sector_short,
                "Aggregation_level": level,
                "Parent_label": parent_label_str,
                "Label": label_str,
                "Type": sheet_type.lower(),
                "Indicator": indicator,
                "Value": val,
                "Code": code
            }
            rows_out.append(rowd)
    logger.info("Extracted %d rows from sheet %s", len(rows_out), sheet.title)
    return rows_out


def process_workbook(path: Path, country: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    out: List[Dict[str, Any]] = []
    for sheet in wb.worksheets:
        title = sheet.title.strip()
        lower = title.lower()
        for suffix in ("fec", "ued", "emi"):
            if not lower.startswith("ind_summary"):
                if lower.endswith(suffix):
                    out.extend(process_sheet(sheet, country, suffix))
                    break
    return out


def find_industry_workbooks(country_dir: Path) -> List[Path]:
    # Look for files matching *Industry*.xlsx (case insensitive)
    matches = [p for p in country_dir.glob("*.xlsx") if "industry" in p.name.lower()]
    return matches
#
# def get_fuel_type(df: pd.DataFrame) -> pd.DataFrame:
#     """Infer fuel type from Indicator column using simple heuristics."""
#
#     #Fuel types (directly inferred from level_4 or level_5)
#     possible_fuels = {'LPG', 'Diesel oil and liquid biofuels', 'Fuel oil',
#        'Natural gas and biogas', 'Solids', 'Refinery gas',
#        'Other liquids', 'Derived gases', 'Biomass and waste',
#        'Distributed steam', 'Diesel oil and liquid biofuels', 'Natural gas and biogas',
#        'Solar and geothermal', 'Ambient heat', 'Electricity', 'Solids',
#        'Fuel oil', 'Derived gases', 'Coke', 'Diesel oil', 'Natural gas', 'Naphtha'}
#
#     electricity_driven_activities = {'Lighting', 'Air compressors', 'Motor drives', 'Fans and pumps'}
#
#     def extract_fuel(row):
#         for col in ["level_3"]:
#             if row[col] in electricity_driven_activities:
#                 return "Electricity"
#
#         for col in ["level_3", "level_4"]:
#             value = row.get(col)
#             if isinstance(value, str):
#                 value_lower = value.lower()
#                 if any(k in value_lower for k in ["electr", "Microwave", "grinding"]):
#                     return "Electricity"
#
#         for col in ["level_3", "level_4"]:
#             value = row.get(col)
#             if isinstance(value, str):
#                 value_lower = value.lower()
#                 if any(k in value_lower for k in ["Natural gas and biogas"]):
#                     return "Natural gas and biogas"
#
#         for col in ["level_4", "level_5"]:
#             val = row[col]
#             if pd.notna(val) and val in possible_fuels:
#                 return val
#         return "Other"
#
#     df["fuel_type"] = df.apply(extract_fuel, axis=1)
#     return df





input_dir = Path("C:/Users/jwiegner/PycharmProjects/CarnotCostTargeting/data/JRC-IDEES2023")
out_dir = Path("C:/Users/jwiegner/PycharmProjects/CarnotCostTargeting/data/JRC-IDEES2023")
countries = discover_countries(input_dir)

for cdir in countries:
    all_rows: List[Dict[str, Any]] = []
    country_code = cdir.name
    wbs = find_industry_workbooks(cdir)
    for wb in wbs:
        logger.info("Processing workbook %s", wb)
        try:
            rows = process_workbook(wb, country_code)
            all_rows.extend(rows)
        except Exception as e:
            logger.exception("Failed to process workbook %s: %s", wb, e)

    df = pd.DataFrame(all_rows)
    df = df[df["Value"].notna()]
    df = df[df["Code"].notna()]

    df = df.join(df["Code"].str.split(".", expand=True).rename(
        columns={0: "Variable", 1: "Unit", 2: "MS_code", 3: "Sector", 4: "Subsector", 5: "Process", 6: "End_use",
                 7: "Fuel"}
    ))



    # levels = (
    #     df["Label"]
    #     .str.split(">", expand=True)
    #     .apply(lambda c: c.str.strip())
    # )
    #
    # levels.columns = [f"level_{i + 1}" for i in range(levels.shape[1])]
    # df = pd.concat([df, levels], axis=1)
    # df = df.rename(columns={"level_1": "variable_type"})
    # df = df.rename(columns={"level_2": "sub_sector"})



    out_file = out_dir / f"jrc_idees_industry_long_{country_code}.csv"
    df.to_csv(out_file, index=False)
    logger.info("Wrote %d rows to %s", len(df), out_file)
